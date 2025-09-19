"""
測試新的Excel輸出格式

驗證垂直合併的分層chunks輸出是否正常工作。
"""

import sys
import logging
from pathlib import Path

# 添加路徑到 Python 路徑
current_dir = Path(__file__).parent
project_root = current_dir.parent.parent
sys.path.insert(0, str(project_root))

from service.chunk.hierarchical_splitter import HierarchicalChunkSplitter
from service.chunk.hierarchical_models import GroupingAnalysis

# 設定日誌
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


def test_excel_output_format():
    """測試Excel輸出格式"""
    logger.info("=== 測試Excel輸出格式 ===")
    
    try:
        # 創建分層分割器
        splitter = HierarchicalChunkSplitter(
            parent_chunk_size=1000,
            child_chunk_size=300,
            child_chunk_overlap=50
        )
        
        # 測試文件
        test_file = project_root / "raw_docs" / "理賠審核原則.xlsx"
        
        if not test_file.exists():
            logger.warning(f"測試文件不存在: {test_file}")
            return False
        
        # 進行分層分割
        logger.info("執行分層分割...")
        result = splitter.split_hierarchically(
            input_data=str(test_file),
            output_excel=True,
            output_path=str(project_root / "service" / "output" / "test_merged_output.xlsx")
        )
        
        # 檢查結果
        logger.info(f"分割結果:")
        logger.info(f"- 父chunks: {len(result.parent_chunks)}")
        logger.info(f"- 子chunks: {len(result.child_chunks)}")
        
        # 檢查Excel文件是否創建
        output_file = project_root / "service" / "output" / "test_merged_output.xlsx"
        if output_file.exists():
            logger.info(f"✓ Excel文件已創建: {output_file}")
            logger.info(f"文件大小: {output_file.stat().st_size} bytes")
            
            # 檢查工作表
            try:
                from openpyxl import load_workbook
                wb = load_workbook(str(output_file))
                sheet_names = wb.sheetnames
                logger.info(f"工作表: {sheet_names}")
                
                # 檢查主要工作表
                if "分層Chunks" in sheet_names:
                    main_sheet = wb["分層Chunks"]
                    logger.info(f"主要工作表行數: {main_sheet.max_row}")
                    logger.info(f"主要工作表列數: {main_sheet.max_column}")
                    
                    # 檢查標題行
                    headers = []
                    for col in range(1, main_sheet.max_column + 1):
                        headers.append(main_sheet.cell(row=1, column=col).value)
                    logger.info(f"標題行: {headers}")
                    
                    # 檢查前幾行數據
                    logger.info("前5行數據:")
                    for row in range(2, min(7, main_sheet.max_row + 1)):
                        row_data = []
                        for col in range(1, main_sheet.max_column + 1):
                            cell_value = main_sheet.cell(row=row, column=col).value
                            row_data.append(str(cell_value)[:20] if cell_value else "")
                        logger.info(f"  行{row}: {row_data}")
                
                return True
                
            except Exception as e:
                logger.error(f"讀取Excel文件失敗: {e}")
                return False
        else:
            logger.error("Excel文件未創建")
            return False
            
    except Exception as e:
        logger.error(f"測試失敗: {e}")
        return False


def test_data_structure():
    """測試資料結構"""
    logger.info("=== 測試資料結構 ===")
    
    try:
        # 創建測試資料
        parent_data = [
            {
                'chunk_id': 'parent_001',
                'index': 0,
                'size': 500,
                'has_tables': True,
                'table_count': 1,
                'header_level': 'Header 1',
                'header_text': '測試標題1',
                'page_number': 1,
                'content': '這是父chunk 1的內容...'
            },
            {
                'chunk_id': 'parent_002',
                'index': 1,
                'size': 800,
                'has_tables': False,
                'table_count': 0,
                'header_level': 'Header 2',
                'header_text': '測試標題2',
                'page_number': 2,
                'content': '這是父chunk 2的內容...'
            }
        ]
        
        child_data = [
            {
                'chunk_id': 'child_001',
                'parent_chunk_id': 'parent_001',
                'child_index': 0,
                'size': 300,
                'is_table_chunk': True,
                'parent_header': '測試標題1',
                'page_number': 1,
                'content': '這是子chunk 1的內容...'
            },
            {
                'chunk_id': 'child_002',
                'parent_chunk_id': 'parent_001',
                'child_index': 1,
                'size': 200,
                'is_table_chunk': False,
                'parent_header': '測試標題1',
                'page_number': 1,
                'content': '這是子chunk 2的內容...'
            },
            {
                'chunk_id': 'child_003',
                'parent_chunk_id': 'parent_002',
                'child_index': 0,
                'size': 400,
                'is_table_chunk': False,
                'parent_header': '測試標題2',
                'page_number': 2,
                'content': '這是子chunk 3的內容...'
            }
        ]
        
        # 創建分組分析
        grouping_analysis = GroupingAnalysis(
            total_parent_chunks=2,
            total_child_chunks=3,
            avg_children_per_parent=1.5,
            parent_size_stats={'min': 500, 'max': 800, 'avg': 650, 'median': 650},
            child_size_stats={'min': 200, 'max': 400, 'avg': 300, 'median': 300},
            table_handling_stats={
                'total_table_chunks': 1,
                'total_regular_chunks': 2,
                'table_chunk_ratio': 0.33,
                'avg_table_size': 300,
                'largest_table_size': 300,
                'table_fragmentation_count': 0
            },
            grouping_efficiency=0.85,
            size_distribution={'0-200': 0, '200-400': 3, '400-600': 0, '600-800': 0, '800-1000': 0, '1000+': 0}
        )
        
        # 測試Excel導出
        from service.chunk.excel_exporter import ExcelExporter
        
        exporter = ExcelExporter()
        output_path = str(project_root / "service" / "output" / "test_data_structure.xlsx")
        
        exporter.export_hierarchical_chunks_to_excel(
            parent_data, child_data, grouping_analysis, output_path
        )
        
        # 檢查輸出文件
        output_file = Path(output_path)
        if output_file.exists():
            logger.info(f"✓ 測試Excel文件已創建: {output_file}")
            
            # 讀取並檢查內容
            from openpyxl import load_workbook
            wb = load_workbook(str(output_file))
            
            if "分層Chunks" in wb.sheetnames:
                main_sheet = wb["分層Chunks"]
                logger.info(f"主要工作表行數: {main_sheet.max_row}")
                
                # 檢查層級分佈
                parent_rows = 0
                child_rows = 0
                
                for row in range(2, main_sheet.max_row + 1):
                    level = main_sheet.cell(row=row, column=1).value
                    if level == "父層":
                        parent_rows += 1
                    elif level == "子層":
                        child_rows += 1
                
                logger.info(f"父層行數: {parent_rows}")
                logger.info(f"子層行數: {child_rows}")
                
                if parent_rows == 2 and child_rows == 3:
                    logger.info("✓ 資料結構正確")
                    return True
                else:
                    logger.error("✗ 資料結構不正確")
                    return False
            else:
                logger.error("✗ 找不到主要工作表")
                return False
        else:
            logger.error("✗ 測試文件未創建")
            return False
            
    except Exception as e:
        logger.error(f"資料結構測試失敗: {e}")
        return False


def main():
    """主測試程式"""
    logger.info("開始Excel輸出格式測試...")
    
    tests = [
        ("資料結構測試", test_data_structure),
        ("實際文件測試", test_excel_output_format)
    ]
    
    results = []
    
    for test_name, test_func in tests:
        logger.info(f"\n執行測試: {test_name}")
        try:
            result = test_func()
            results.append((test_name, result))
            if result:
                logger.info(f"✓ {test_name} 通過")
            else:
                logger.error(f"✗ {test_name} 失敗")
        except Exception as e:
            logger.error(f"✗ {test_name} 異常: {e}")
            results.append((test_name, False))
    
    # 顯示測試結果摘要
    logger.info("\n" + "="*50)
    logger.info("測試結果摘要:")
    
    passed = 0
    total = len(results)
    
    for test_name, result in results:
        status = "✓ 通過" if result else "✗ 失敗"
        logger.info(f"- {test_name}: {status}")
        if result:
            passed += 1
    
    logger.info(f"\n總計: {passed}/{total} 測試通過")
    
    if passed == total:
        logger.info("🎉 所有測試通過！Excel輸出格式已準備就緒。")
    else:
        logger.warning("⚠ 部分測試失敗，請檢查相關功能。")


if __name__ == "__main__":
    main()
