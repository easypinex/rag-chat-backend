"""
Excel輸出格式範例

展示新的垂直合併Excel輸出格式的使用方法。
"""

import sys
import logging
from pathlib import Path

# 添加路徑到 Python 路徑
current_dir = Path(__file__).parent
project_root = current_dir.parent.parent.parent
sys.path.insert(0, str(project_root))

from service.chunk.hierarchical_splitter import HierarchicalChunkSplitter
from service.chunk.excel_exporter import ExcelExporter
from service.chunk.hierarchical_models import GroupingAnalysis

# 設定日誌
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


def example_merged_excel_output():
    """範例：垂直合併的Excel輸出"""
    logger.info("=== 垂直合併Excel輸出範例 ===")
    
    try:
        # 創建分層分割器
        splitter = HierarchicalChunkSplitter(
            parent_chunk_size=1500,
            child_chunk_size=350,
            child_chunk_overlap=50
        )
        
        # 測試文件
        test_file = project_root / "raw_docs" / "理賠審核原則.xlsx"
        
        if not test_file.exists():
            logger.warning(f"測試文件不存在: {test_file}")
            return None
        
        # 進行分層分割
        logger.info("執行分層分割...")
        result = splitter.split_hierarchically(
            input_data=str(test_file),
            output_excel=True,
            output_path=str(project_root / "service" / "output" / "merged_excel_example.xlsx")
        )
        
        # 顯示結果信息
        logger.info(f"分割結果:")
        logger.info(f"- 父chunks: {len(result.parent_chunks)}")
        logger.info(f"- 子chunks: {len(result.child_chunks)}")
        logger.info(f"- 分組效率: {result.grouping_analysis.grouping_efficiency:.2%}")
        
        # 檢查Excel文件
        output_file = project_root / "service" / "output" / "merged_excel_example.xlsx"
        if output_file.exists():
            logger.info(f"✓ Excel文件已創建: {output_file}")
            
            # 讀取Excel文件並顯示結構
            from openpyxl import load_workbook
            wb = load_workbook(str(output_file))
            
            logger.info(f"工作表: {wb.sheetnames}")
            
            # 檢查主要工作表
            if "分層Chunks" in wb.sheetnames:
                main_sheet = wb["分層Chunks"]
                logger.info(f"主要工作表行數: {main_sheet.max_row}")
                
                # 顯示前幾行數據結構
                logger.info("前10行數據結構:")
                for row in range(1, min(11, main_sheet.max_row + 1)):
                    row_data = []
                    for col in range(1, min(6, main_sheet.max_column + 1)):  # 只顯示前5列
                        cell_value = main_sheet.cell(row=row, column=col).value
                        row_data.append(str(cell_value)[:15] if cell_value else "")
                    logger.info(f"  行{row}: {row_data}")
                
                # 統計層級分佈
                parent_count = 0
                child_count = 0
                
                for row in range(2, main_sheet.max_row + 1):
                    level = main_sheet.cell(row=row, column=1).value
                    if level == "父層":
                        parent_count += 1
                    elif level == "子層":
                        child_count += 1
                
                logger.info(f"Excel中的層級分佈:")
                logger.info(f"- 父層行數: {parent_count}")
                logger.info(f"- 子層行數: {child_count}")
                
                # 檢查父子關係
                logger.info("父子關係檢查:")
                for row in range(2, min(10, main_sheet.max_row + 1)):
                    level = main_sheet.cell(row=row, column=1).value
                    chunk_id = main_sheet.cell(row=row, column=2).value
                    parent_id = main_sheet.cell(row=row, column=3).value
                    
                    if level == "父層":
                        logger.info(f"  父chunk: {chunk_id} (無父ID)")
                    elif level == "子層":
                        logger.info(f"  子chunk: {chunk_id} -> 父chunk: {parent_id}")
        
        return result
        
    except Exception as e:
        logger.error(f"範例執行失敗: {e}")
        return None


def example_custom_data_structure():
    """範例：自定義資料結構的Excel輸出"""
    logger.info("=== 自定義資料結構範例 ===")
    
    try:
        # 創建測試資料
        parent_data = [
            {
                'chunk_id': 'parent_001',
                'index': 0,
                'size': 1200,
                'has_tables': True,
                'table_count': 2,
                'header_level': 'Header 1',
                'header_text': '第一章 基本概念',
                'page_number': 1,
                'content': '這是第一章的內容，包含基本概念和定義...'
            },
            {
                'chunk_id': 'parent_002',
                'index': 1,
                'size': 1800,
                'has_tables': False,
                'table_count': 0,
                'header_level': 'Header 2',
                'header_text': '第二章 實施方法',
                'page_number': 2,
                'content': '這是第二章的內容，詳細說明實施方法和步驟...'
            }
        ]
        
        child_data = [
            {
                'chunk_id': 'child_001',
                'parent_chunk_id': 'parent_001',
                'child_index': 0,
                'size': 350,
                'is_table_chunk': True,
                'parent_header': '第一章 基本概念',
                'page_number': 1,
                'content': '基本概念的第一部分內容，包含重要定義...'
            },
            {
                'chunk_id': 'child_002',
                'parent_chunk_id': 'parent_001',
                'child_index': 1,
                'size': 300,
                'is_table_chunk': True,
                'parent_header': '第一章 基本概念',
                'page_number': 1,
                'content': '基本概念的第二部分內容，包含範例說明...'
            },
            {
                'chunk_id': 'child_003',
                'parent_chunk_id': 'parent_001',
                'child_index': 2,
                'size': 250,
                'is_table_chunk': False,
                'parent_header': '第一章 基本概念',
                'page_number': 1,
                'content': '基本概念的第三部分內容，包含總結...'
            },
            {
                'chunk_id': 'child_004',
                'parent_chunk_id': 'parent_002',
                'child_index': 0,
                'size': 400,
                'is_table_chunk': False,
                'parent_header': '第二章 實施方法',
                'page_number': 2,
                'content': '實施方法的第一部分內容，包含基本步驟...'
            },
            {
                'chunk_id': 'child_005',
                'parent_chunk_id': 'parent_002',
                'child_index': 1,
                'size': 350,
                'is_table_chunk': False,
                'parent_header': '第二章 實施方法',
                'page_number': 2,
                'content': '實施方法的第二部分內容，包含詳細說明...'
            }
        ]
        
        # 創建分組分析
        grouping_analysis = GroupingAnalysis(
            total_parent_chunks=2,
            total_child_chunks=5,
            avg_children_per_parent=2.5,
            parent_size_stats={'min': 1200, 'max': 1800, 'avg': 1500, 'median': 1500},
            child_size_stats={'min': 250, 'max': 400, 'avg': 330, 'median': 350},
            table_handling_stats={
                'total_table_chunks': 2,
                'total_regular_chunks': 3,
                'table_chunk_ratio': 0.4,
                'avg_table_size': 325,
                'largest_table_size': 350,
                'table_fragmentation_count': 1
            },
            grouping_efficiency=0.92,
            size_distribution={'0-200': 0, '200-400': 5, '400-600': 0, '600-800': 0, '800-1000': 0, '1000+': 0}
        )
        
        # 導出到Excel
        exporter = ExcelExporter()
        output_path = str(project_root / "service" / "output" / "custom_data_example.xlsx")
        
        exporter.export_hierarchical_chunks_to_excel(
            parent_data, child_data, grouping_analysis, output_path
        )
        
        logger.info(f"✓ 自定義資料Excel文件已創建: {output_path}")
        
        # 讀取並顯示結果
        from openpyxl import load_workbook
        wb = load_workbook(str(output_path))
        
        if "分層Chunks" in wb.sheetnames:
            main_sheet = wb["分層Chunks"]
            logger.info(f"工作表行數: {main_sheet.max_row}")
            
            # 顯示完整的資料結構
            logger.info("完整資料結構:")
            for row in range(1, main_sheet.max_row + 1):
                level = main_sheet.cell(row=row, column=1).value
                chunk_id = main_sheet.cell(row=row, column=2).value
                parent_id = main_sheet.cell(row=row, column=3).value
                size = main_sheet.cell(row=row, column=5).value
                is_table = main_sheet.cell(row=row, column=6).value
                header = main_sheet.cell(row=row, column=8).value
                
                if level == "父層":
                    logger.info(f"  {level}: {chunk_id} (大小: {size}, 表格: {is_table}, 標題: {header})")
                elif level == "子層":
                    logger.info(f"    {level}: {chunk_id} -> {parent_id} (大小: {size}, 表格: {is_table})")
        
        return True
        
    except Exception as e:
        logger.error(f"自定義資料範例失敗: {e}")
        return False


def main():
    """主程式"""
    logger.info("開始Excel輸出格式範例...")
    
    # 範例1: 實際文件的分層分割
    logger.info("\n" + "="*50)
    result1 = example_merged_excel_output()
    
    # 範例2: 自定義資料結構
    logger.info("\n" + "="*50)
    result2 = example_custom_data_structure()
    
    logger.info("\n所有範例執行完成！")
    logger.info("請檢查 service/output/ 目錄中的Excel文件。")


if __name__ == "__main__":
    main()
