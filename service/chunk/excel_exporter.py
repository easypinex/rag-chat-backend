"""
Excel 導出器

將分割後的 chunks 導出到 Excel 文件，支援合併單元格功能。
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path
from typing import List, Dict, Any, Optional
import logging

from langchain_core.documents import Document
from .hierarchical_models import GroupingAnalysis

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Excel 導出器"""
    
    def __init__(self):
        """初始化導出器"""
        self.workbook = None
        self.worksheet = None
    
    def export_chunks_to_excel(self, 
                              chunks: List[Document], 
                              original_content: str,
                              output_path: str,
                              normalized_content: Optional[str] = None):
        """
        將 chunks 導出到 Excel 文件
        
        Args:
            chunks: 分割後的文檔列表
            original_content: 原始 Markdown 內容
            output_path: 輸出文件路徑
            normalized_content: 正規化後的內容（可選）
        """
        # 確保輸出目錄存在
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        
        # 創建工作簿
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Markdown Chunks"
        
        # 設置標題行
        self._setup_headers()
        
        # 填充數據
        self._fill_data(chunks, original_content, normalized_content)
        
        # 設置格式
        self._format_worksheet()
        
        # 保存文件
        self.workbook.save(output_path)
        logger.info(f"Excel file saved to: {output_path}")
    
    def export_chunks_to_excel_with_page_content(self, chunks: List[Document], page_chunks_info: List[Dict], output_path: str):
        """
        將 chunks 導出到 Excel 文件，使用按頁面分割的原始和正規化內容
        
        Args:
            chunks: 分割後的文檔列表
            page_chunks_info: 頁面 chunks 信息列表
            output_path: 輸出文件路徑
        """
        # 確保輸出目錄存在
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        
        # 創建工作簿
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Markdown Chunks"
        
        # 設置標題行
        self._setup_headers()
        
        # 填充數據
        self._fill_data_with_page_content(chunks, page_chunks_info)
        
        # 設置格式
        self._format_worksheet()
        
        # 保存文件
        self.workbook.save(output_path)
        logger.info(f"Excel file saved to: {output_path}")
    
    def _setup_headers(self):
        """設置標題行"""
        headers = [
            "原始內容 (A欄)",
            "正規化後內容 (B欄)",
            "分割後的 Chunk (C欄)", 
            "Chunk 編號",
            "Chunk 長度",
            "包含標題",
            "是否為表格",
            # 展開的 metadata 欄位
            "檔名",
            "檔案類型",
            "來源路徑",
            "轉換器",
            "總頁數",
            "總表格數",
            "頁碼",
            "頁面標題",
            "標題級數",
            "表格合併數",
            "完整元數據"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=1, column=col, value=header)
            cell.font = cell.font.copy(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    def _fill_data(self, chunks: List[Document], original_content: str, normalized_content: Optional[str] = None):
        """填充數據到工作表"""
        current_row = 2
        
        # 如果沒有提供正規化內容，使用原始內容
        if normalized_content is None:
            normalized_content = original_content
        
        for i, chunk in enumerate(chunks, 1):
            # A欄：原始內容（會根據 chunks 數量進行合併）
            self.worksheet.cell(row=current_row, column=1, value=original_content)
            
            # B欄：正規化後內容（會根據 chunks 數量進行合併）
            self.worksheet.cell(row=current_row, column=2, value=normalized_content)
            
            # C欄：分割後的 Chunk
            self.worksheet.cell(row=current_row, column=3, value=chunk.page_content)
            
            # D欄：Chunk 編號（使用全局編號）
            global_chunk_number = chunk.metadata.get('global_chunk_number', i)
            self.worksheet.cell(row=current_row, column=4, value=global_chunk_number)
            
            # E欄：Chunk 長度
            self.worksheet.cell(row=current_row, column=5, value=len(chunk.page_content))
            
            # F欄：包含標題
            has_headers = any(header in chunk.page_content for header in ['#', '##', '###', '####'])
            self.worksheet.cell(row=current_row, column=6, value="是" if has_headers else "否")
            
            # G欄：是否為表格
            is_table = chunk.metadata.get('is_table', False)
            self.worksheet.cell(row=current_row, column=7, value="是" if is_table else "否")
            
            # 展開的 metadata 欄位 (H欄開始)
            # H欄：檔名
            self.worksheet.cell(row=current_row, column=8, value=chunk.metadata.get('file_name', ''))
            
            # I欄：檔案類型
            self.worksheet.cell(row=current_row, column=9, value=chunk.metadata.get('file_type', ''))
            
            # J欄：來源路徑
            self.worksheet.cell(row=current_row, column=10, value=chunk.metadata.get('source', ''))
            
            # K欄：轉換器
            self.worksheet.cell(row=current_row, column=11, value=chunk.metadata.get('converter_used', ''))
            
            # L欄：總頁數
            self.worksheet.cell(row=current_row, column=12, value=chunk.metadata.get('total_pages', ''))
            
            # M欄：總表格數
            self.worksheet.cell(row=current_row, column=13, value=chunk.metadata.get('total_tables', ''))
            
            # N欄：頁碼
            self.worksheet.cell(row=current_row, column=14, value=chunk.metadata.get('page_number', ''))
            
            # O欄：頁面標題
            self.worksheet.cell(row=current_row, column=15, value=chunk.metadata.get('page_title', ''))
            
            # P欄：標題級數（合併四個標題欄位）
            header_level = self._get_header_level(chunk.metadata)
            self.worksheet.cell(row=current_row, column=16, value=header_level)
            
            # Q欄：表格合併數
            self.worksheet.cell(row=current_row, column=17, value=chunk.metadata.get('table_chunks_merged', ''))
            
            # R欄：完整元數據
            metadata_str = self._format_metadata(chunk.metadata)
            self.worksheet.cell(row=current_row, column=18, value=metadata_str)
            
            current_row += 1
        
        # 合併 A欄 和 B欄 的單元格（因為原始內容和正規化內容相同）
        if len(chunks) > 1:
            self.worksheet.merge_cells(f'A2:A{current_row - 1}')
            self.worksheet.merge_cells(f'B2:B{current_row - 1}')
    
    def _fill_data_with_page_content(self, chunks: List[Document], page_chunks_info: List[Dict]):
        """填充數據到工作表，使用按頁面分割的內容"""
        current_row = 2
        
        for i, chunk in enumerate(chunks, 1):
            # 獲取頁面級別的原始和正規化內容
            page_number = chunk.metadata.get('page_number')
            
            # 找到對應的頁面信息
            page_info = None
            for page_data in page_chunks_info:
                if page_data['page_number'] == page_number:
                    page_info = page_data
                    break
            
            # 使用頁面內容或 chunk 內容作為備用
            if page_info:
                page_original_content = page_info['original_content']
                page_normalized_content = page_info['normalized_content']
            else:
                page_original_content = chunk.page_content
                page_normalized_content = chunk.page_content
            
            # A欄：頁面原始內容
            self.worksheet.cell(row=current_row, column=1, value=page_original_content)
            
            # B欄：頁面正規化內容
            self.worksheet.cell(row=current_row, column=2, value=page_normalized_content)
            
            # C欄：分割後的 Chunk
            self.worksheet.cell(row=current_row, column=3, value=chunk.page_content)
            
            # D欄：Chunk 編號（使用全局編號）
            global_chunk_number = chunk.metadata.get('global_chunk_number', i)
            self.worksheet.cell(row=current_row, column=4, value=global_chunk_number)
            
            # E欄：Chunk 長度
            self.worksheet.cell(row=current_row, column=5, value=len(chunk.page_content))
            
            # F欄：包含標題
            has_headers = any(header in chunk.page_content for header in ['#', '##', '###', '####'])
            self.worksheet.cell(row=current_row, column=6, value="是" if has_headers else "否")
            
            # G欄：是否為表格
            is_table = chunk.metadata.get('is_table', False)
            self.worksheet.cell(row=current_row, column=7, value="是" if is_table else "否")
            
            # 展開的 metadata 欄位 (H欄開始)
            # H欄：檔名
            self.worksheet.cell(row=current_row, column=8, value=chunk.metadata.get('file_name', ''))
            
            # I欄：檔案類型
            self.worksheet.cell(row=current_row, column=9, value=chunk.metadata.get('file_type', ''))
            
            # J欄：來源路徑
            self.worksheet.cell(row=current_row, column=10, value=chunk.metadata.get('source', ''))
            
            # K欄：轉換器
            self.worksheet.cell(row=current_row, column=11, value=chunk.metadata.get('converter_used', ''))
            
            # L欄：總頁數
            self.worksheet.cell(row=current_row, column=12, value=chunk.metadata.get('total_pages', ''))
            
            # M欄：總表格數
            self.worksheet.cell(row=current_row, column=13, value=chunk.metadata.get('total_tables', ''))
            
            # N欄：頁碼
            self.worksheet.cell(row=current_row, column=14, value=chunk.metadata.get('page_number', ''))
            
            # O欄：頁面標題
            self.worksheet.cell(row=current_row, column=15, value=chunk.metadata.get('page_title', ''))
            
            # P欄：標題級數（合併四個標題欄位）
            header_level = self._get_header_level(chunk.metadata)
            self.worksheet.cell(row=current_row, column=16, value=header_level)
            
            # Q欄：表格合併數
            self.worksheet.cell(row=current_row, column=17, value=chunk.metadata.get('table_chunks_merged', ''))
            
            # R欄：完整元數據
            metadata_str = self._format_metadata(chunk.metadata)
            self.worksheet.cell(row=current_row, column=18, value=metadata_str)
            
            current_row += 1
        
        # 按頁碼分組合併 A欄 和 B欄 的單元格
        self._merge_cells_by_page(chunks)
    
    def _merge_cells_by_page(self, chunks: List[Document]):
        """按頁碼分組合併單元格"""
        # 按頁碼分組
        page_groups = {}
        for i, chunk in enumerate(chunks, 2):  # 從第2行開始
            page_number = chunk.metadata.get('page_number')
            # 如果沒有頁碼，使用 'no_page' 作為分組鍵
            if page_number is None:
                page_number = 'no_page'
            if page_number not in page_groups:
                page_groups[page_number] = []
            page_groups[page_number].append(i)
        
        # 為每個頁碼組合併單元格
        for page_number, rows in page_groups.items():
            if len(rows) > 1:
                start_row = min(rows)
                end_row = max(rows)
                self.worksheet.merge_cells(f'A{start_row}:A{end_row}')
                self.worksheet.merge_cells(f'B{start_row}:B{end_row}')
    
    def _format_metadata(self, metadata: Dict[str, Any]) -> str:
        """格式化元數據為字符串"""
        if not metadata:
            return "無"
        
        # 優先顯示重要信息
        important_keys = ['file_name', 'page_number', 'Header 1', 'Header 2', 'Header 3', 'Header 4', 'is_table']
        formatted_items = []
        
        # 先添加重要信息
        for key in important_keys:
            if key in metadata:
                value = metadata[key]
                if isinstance(value, (str, int, float, bool)):
                    formatted_items.append(f"{key}: {value}")
                else:
                    formatted_items.append(f"{key}: {str(value)[:30]}...")
        
        # 再添加其他信息
        for key, value in metadata.items():
            if key not in important_keys:
                if isinstance(value, (str, int, float, bool)):
                    formatted_items.append(f"{key}: {value}")
                else:
                    formatted_items.append(f"{key}: {str(value)[:30]}...")
        
        return "; ".join(formatted_items)
    
    def _get_header_level(self, metadata: Dict[str, Any]) -> str:
        """獲取標題級數（一、二、三、四）"""
        if 'Header 4' in metadata and metadata['Header 4']:
            return '四'
        elif 'Header 3' in metadata and metadata['Header 3']:
            return '三'
        elif 'Header 2' in metadata and metadata['Header 2']:
            return '二'
        elif 'Header 1' in metadata and metadata['Header 1']:
            return '一'
        else:
            return ''
    
    def _format_worksheet(self):
        """設置工作表格式"""
        # 設置列寬
        column_widths = {
            'A': 30,  # 原始內容
            'B': 30,  # 正規化後內容
            'C': 50,  # Chunk 內容
            'D': 10,  # 編號
            'E': 12,  # 長度
            'F': 12,  # 包含標題
            'G': 12,  # 是否表格
            'H': 25,  # 檔名
            'I': 12,  # 檔案類型
            'J': 30,  # 來源路徑
            'K': 12,  # 轉換器
            'L': 10,  # 總頁數
            'M': 12,  # 總表格數
            'N': 10,  # 頁碼
            'O': 20,  # 頁面標題
            'P': 12,  # 標題級數
            'Q': 12,  # 表格合併數
            'R': 30   # 完整元數據
        }
        
        for col, width in column_widths.items():
            self.worksheet.column_dimensions[col].width = width
        
        # 設置邊框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 為所有數據單元格設置邊框
        for row in self.worksheet.iter_rows(min_row=1, max_row=self.worksheet.max_row):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        
        # 設置 A欄 和 B欄 的對齊方式（合併的單元格）
        for row in range(2, self.worksheet.max_row + 1):
            # A欄：原始內容
            cell_a = self.worksheet.cell(row=row, column=1)
            cell_a.alignment = Alignment(vertical='top', horizontal='left', wrap_text=True)
            
            # B欄：正規化後內容
            cell_b = self.worksheet.cell(row=row, column=2)
            cell_b.alignment = Alignment(vertical='top', horizontal='left', wrap_text=True)
    
    def create_summary_sheet(self, chunks: List[Document], output_path: str):
        """創建統計摘要工作表"""
        if not self.workbook:
            self.workbook = Workbook()
        
        # 創建摘要工作表
        summary_sheet = self.workbook.create_sheet("統計摘要")
        
        # 統計信息
        total_chunks = len(chunks)
        total_length = sum(len(chunk.page_content) for chunk in chunks)
        avg_length = total_length / total_chunks if total_chunks > 0 else 0
        table_chunks = sum(1 for chunk in chunks if chunk.metadata.get('is_table', False))
        
        # 填充統計信息
        stats = [
            ("總 Chunk 數量", total_chunks),
            ("總字符數", total_length),
            ("平均 Chunk 長度", round(avg_length, 2)),
            ("表格 Chunk 數量", table_chunks),
            ("一般 Chunk 數量", total_chunks - table_chunks),
            ("表格比例", f"{round(table_chunks / total_chunks * 100, 2)}%" if total_chunks > 0 else "0%")
        ]
        
        for i, (label, value) in enumerate(stats, 1):
            summary_sheet.cell(row=i, column=1, value=label).font = summary_sheet.cell(row=i, column=1).font.copy(bold=True)
            summary_sheet.cell(row=i, column=2, value=value)
        
        # 保存文件
        self.workbook.save(output_path)
        logger.info(f"Summary sheet added to: {output_path}")
    
    def export_hierarchical_chunks_to_excel(self, 
                                           parent_data: List[Dict[str, Any]], 
                                           child_data: List[Dict[str, Any]], 
                                           grouping_analysis: GroupingAnalysis,
                                           output_path: str):
        """
        導出分層chunks到Excel文件 - 使用垂直合併方式
        
        Args:
            parent_data: 父chunks數據
            child_data: 子chunks數據
            grouping_analysis: 分組分析結果
            output_path: 輸出路徑
        """
        # 確保輸出目錄存在
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        
        # 創建工作簿
        self.workbook = Workbook()
        
        # 移除默認工作表
        if "Sheet" in self.workbook.sheetnames:
            self.workbook.remove(self.workbook["Sheet"])
        
        # 1. 創建主要的分層chunks工作表（垂直合併）
        self._create_hierarchical_merged_sheet(parent_data, child_data)
        
        # 2. 創建只有父Chunk的工作表
        self._create_parent_only_sheet(parent_data)
        
        # 3. 創建只有子Chunk的工作表
        self._create_child_only_sheet(child_data)
        
        # 4. 創建分組分析工作表
        self._create_grouping_analysis_sheet(grouping_analysis)
        
        # 5. 創建統計摘要工作表
        self._create_hierarchical_summary_sheet(parent_data, child_data, grouping_analysis)
        
        # 保存文件
        self.workbook.save(output_path)
        logger.info(f"Hierarchical chunks exported to: {output_path}")
    
    
    def _create_grouping_analysis_sheet(self, grouping_analysis: GroupingAnalysis):
        """創建分組分析工作表"""
        analysis_sheet = self.workbook.create_sheet("分組分析")
        
        # 基本統計
        basic_stats = [
            ("總父Chunks數量", grouping_analysis.total_parent_chunks),
            ("總子Chunks數量", grouping_analysis.total_child_chunks),
            ("平均每父Chunk的子Chunks數", round(grouping_analysis.avg_children_per_parent, 2)),
            ("分組效率", f"{round(grouping_analysis.grouping_efficiency * 100, 2)}%")
        ]
        
        # 父Chunk大小統計
        parent_size_stats = [
            ("父Chunk最小大小", grouping_analysis.parent_size_stats['min']),
            ("父Chunk最大大小", grouping_analysis.parent_size_stats['max']),
            ("父Chunk平均大小", round(grouping_analysis.parent_size_stats['avg'], 2)),
            ("父Chunk中位數大小", round(grouping_analysis.parent_size_stats['median'], 2))
        ]
        
        # 子Chunk大小統計
        child_size_stats = [
            ("子Chunk最小大小", grouping_analysis.child_size_stats['min']),
            ("子Chunk最大大小", grouping_analysis.child_size_stats['max']),
            ("子Chunk平均大小", round(grouping_analysis.child_size_stats['avg'], 2)),
            ("子Chunk中位數大小", round(grouping_analysis.child_size_stats['median'], 2))
        ]
        
        # 表格處理統計
        table_stats = [
            ("表格Chunks數量", grouping_analysis.table_handling_stats['total_table_chunks']),
            ("一般Chunks數量", grouping_analysis.table_handling_stats['total_regular_chunks']),
            ("表格Chunk比例", f"{round(grouping_analysis.table_handling_stats['table_chunk_ratio'] * 100, 2)}%"),
            ("平均表格大小", round(grouping_analysis.table_handling_stats['avg_table_size'], 2)),
            ("最大表格大小", grouping_analysis.table_handling_stats['largest_table_size']),
            ("表格碎片化數量", grouping_analysis.table_handling_stats['table_fragmentation_count'])
        ]
        
        # 填充數據
        row = 1
        sections = [
            ("基本統計", basic_stats),
            ("父Chunk大小統計", parent_size_stats),
            ("子Chunk大小統計", child_size_stats),
            ("表格處理統計", table_stats)
        ]
        
        for section_title, stats in sections:
            # 添加區段標題
            analysis_sheet.cell(row=row, column=1, value=section_title).font = analysis_sheet.cell(row=row, column=1).font.copy(bold=True)
            row += 1
            
            # 添加統計數據
            for label, value in stats:
                analysis_sheet.cell(row=row, column=1, value=label)
                analysis_sheet.cell(row=row, column=2, value=value)
                row += 1
            
            row += 1  # 空行分隔
        
        # 調整列寬
        analysis_sheet.column_dimensions['A'].width = 30
        analysis_sheet.column_dimensions['B'].width = 20
    
    def _create_hierarchical_summary_sheet(self, 
                                         parent_data: List[Dict[str, Any]], 
                                         child_data: List[Dict[str, Any]], 
                                         grouping_analysis: GroupingAnalysis):
        """創建分層摘要工作表"""
        summary_sheet = self.workbook.create_sheet("分層摘要")
        
        # 計算額外統計
        parent_sizes = [data['size'] for data in parent_data]
        child_sizes = [data['size'] for data in child_data]
        
        # 大小分佈統計
        size_ranges = {
            "0-200": 0, "200-400": 0, "400-600": 0, 
            "600-800": 0, "800-1000": 0, "1000+": 0
        }
        
        for size in child_sizes:
            if size < 200:
                size_ranges["0-200"] += 1
            elif size < 400:
                size_ranges["200-400"] += 1
            elif size < 600:
                size_ranges["400-600"] += 1
            elif size < 800:
                size_ranges["600-800"] += 1
            elif size < 1000:
                size_ranges["800-1000"] += 1
            else:
                size_ranges["1000+"] += 1
        
        # 填充摘要數據
        summary_data = [
            ("分層分割摘要", ""),
            ("", ""),
            ("父Chunks統計", ""),
            ("總數量", len(parent_data)),
            ("平均大小", round(sum(parent_sizes) / len(parent_sizes), 2) if parent_sizes else 0),
            ("", ""),
            ("子Chunks統計", ""),
            ("總數量", len(child_data)),
            ("平均大小", round(sum(child_sizes) / len(child_sizes), 2) if child_sizes else 0),
            ("", ""),
            ("大小分佈", ""),
        ]
        
        for range_name, count in size_ranges.items():
            summary_data.append((f"{range_name}字", count))
        
        summary_data.extend([
            ("", ""),
            ("分組效率", f"{round(grouping_analysis.grouping_efficiency * 100, 2)}%"),
            ("分析時間", grouping_analysis.analysis_timestamp)
        ])
        
        # 填充到工作表
        for row, (label, value) in enumerate(summary_data, 1):
            summary_sheet.cell(row=row, column=1, value=label)
            summary_sheet.cell(row=row, column=2, value=value)
            
            # 設置標題行格式
            if label.endswith("摘要") or label.endswith("統計") or label.endswith("分佈"):
                summary_sheet.cell(row=row, column=1).font = summary_sheet.cell(row=row, column=1).font.copy(bold=True)
        
        # 調整列寬
        summary_sheet.column_dimensions['A'].width = 20
        summary_sheet.column_dimensions['B'].width = 20
    
    def _create_hierarchical_merged_sheet(self, parent_data: List[Dict[str, Any]], child_data: List[Dict[str, Any]]):
        """創建垂直合併的分層chunks工作表 - 按頁面分組，參考chunk_splitter.py"""
        merged_sheet = self.workbook.create_sheet("分層Chunks")
        
        # 設置標題行 - 參考一般輸出格式
        headers = [
            "原文", "正規化", "Parent Chunk", "Sub Chunk", "層級", "Chunk ID", "父Chunk ID", 
            "索引", "大小", "是否表格", "標題層級", "標題文字", "頁碼", "檔名", "檔案類型"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = merged_sheet.cell(row=1, column=col, value=header)
            cell.font = cell.font.copy(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = self._get_header_fill()
        
        # 按頁面分組父chunks
        page_parent_map = {}
        for parent in parent_data:
            page_number = parent.get('page_number', 'no_page')
            if page_number not in page_parent_map:
                page_parent_map[page_number] = []
            page_parent_map[page_number].append(parent)
        
        # 按父chunk分組子chunks
        parent_child_map = {}
        for child in child_data:
            parent_id = child['parent_chunk_id']
            if parent_id not in parent_child_map:
                parent_child_map[parent_id] = []
            parent_child_map[parent_id].append(child)
        
        current_row = 2
        
        # 按頁面遍歷
        for page_number in sorted(page_parent_map.keys()):
            page_parents = page_parent_map[page_number]
            
            # 獲取頁面的原文和正規化內容（從第一個父chunk獲取）
            if page_parents:
                page_original = page_parents[0].get('original_content', '')
                page_normalized = page_parents[0].get('normalized_content', '')
                
                # 如果原文為空，嘗試從content中獲取
                if not page_original:
                    page_original = page_parents[0].get('content', '')
                if not page_normalized:
                    page_normalized = page_parents[0].get('content', '')
            else:
                page_original = ''
                page_normalized = ''
            
            # 遍歷該頁面的每個父chunk
            for parent in page_parents:
                parent_id = parent['chunk_id']
                
                # 添加父chunk行
                merged_sheet.cell(row=current_row, column=1, value=page_original)  # 原文（頁面級別）
                merged_sheet.cell(row=current_row, column=2, value=page_normalized)  # 正規化（頁面級別）
                merged_sheet.cell(row=current_row, column=3, value=parent['content'])  # Parent Chunk
                merged_sheet.cell(row=current_row, column=4, value="")  # Sub Chunk (父chunk沒有)
                merged_sheet.cell(row=current_row, column=5, value="父層").font = merged_sheet.cell(row=current_row, column=5).font.copy(bold=True)
                merged_sheet.cell(row=current_row, column=6, value=parent['chunk_id'])
                merged_sheet.cell(row=current_row, column=7, value="")  # 父chunk沒有父ID
                merged_sheet.cell(row=current_row, column=8, value=parent.get('index', ''))
                merged_sheet.cell(row=current_row, column=9, value=parent['size'])
                merged_sheet.cell(row=current_row, column=10, value=parent['has_tables'])
                merged_sheet.cell(row=current_row, column=11, value=parent.get('header_level', ''))
                merged_sheet.cell(row=current_row, column=12, value=parent.get('header_text', ''))
                merged_sheet.cell(row=current_row, column=13, value=parent.get('page_number', ''))
                merged_sheet.cell(row=current_row, column=14, value=parent.get('file_name', ''))
                merged_sheet.cell(row=current_row, column=15, value=parent.get('file_type', ''))
                
                # 設置父chunk行格式
                self._format_parent_row(merged_sheet, current_row)
                current_row += 1
                
                # 添加對應的子chunks
                children = parent_child_map.get(parent_id, [])
                for child in children:
                    # 子chunk的原文和正規化欄位留空，將通過垂直合併顯示
                    merged_sheet.cell(row=current_row, column=1, value="")  # 原文（留空，將合併）
                    merged_sheet.cell(row=current_row, column=2, value="")  # 正規化（留空，將合併）
                    merged_sheet.cell(row=current_row, column=3, value="")  # Parent Chunk (子chunk沒有)
                    merged_sheet.cell(row=current_row, column=4, value=child['content'])  # Sub Chunk
                    merged_sheet.cell(row=current_row, column=5, value="子層")
                    merged_sheet.cell(row=current_row, column=6, value=child['chunk_id'])
                    merged_sheet.cell(row=current_row, column=7, value=child['parent_chunk_id'])
                    merged_sheet.cell(row=current_row, column=8, value=child.get('child_index', ''))
                    merged_sheet.cell(row=current_row, column=9, value=child['size'])
                    merged_sheet.cell(row=current_row, column=10, value=child['is_table_chunk'])
                    merged_sheet.cell(row=current_row, column=11, value="")  # 子chunk沒有自己的標題層級
                    merged_sheet.cell(row=current_row, column=12, value=child.get('parent_header', ''))
                    merged_sheet.cell(row=current_row, column=13, value=child.get('page_number', ''))
                    merged_sheet.cell(row=current_row, column=14, value=child.get('file_name', ''))
                    merged_sheet.cell(row=current_row, column=15, value=child.get('file_type', ''))
                    
                    # 設置子chunk行格式
                    self._format_child_row(merged_sheet, current_row)
                    current_row += 1
                
                # 添加分隔行
                if current_row <= merged_sheet.max_row + 1:
                    merged_sheet.cell(row=current_row, column=1, value="")
                    current_row += 1
        
        # 調整列寬
        merged_sheet.column_dimensions['A'].width = 50   # 原文
        merged_sheet.column_dimensions['B'].width = 50   # 正規化
        merged_sheet.column_dimensions['C'].width = 50   # Parent Chunk
        merged_sheet.column_dimensions['D'].width = 50   # Sub Chunk
        merged_sheet.column_dimensions['E'].width = 8    # 層級
        merged_sheet.column_dimensions['F'].width = 20   # Chunk ID
        merged_sheet.column_dimensions['G'].width = 20   # 父Chunk ID
        merged_sheet.column_dimensions['H'].width = 8    # 索引
        merged_sheet.column_dimensions['I'].width = 10   # 大小
        merged_sheet.column_dimensions['J'].width = 12   # 是否表格
        merged_sheet.column_dimensions['K'].width = 15   # 標題層級
        merged_sheet.column_dimensions['L'].width = 30   # 標題文字
        merged_sheet.column_dimensions['M'].width = 10   # 頁碼
        merged_sheet.column_dimensions['N'].width = 30   # 檔名
        merged_sheet.column_dimensions['O'].width = 15   # 檔案類型
        
        # 設置所有欄位垂直置頂對齊
        for row in range(2, current_row):
            for col in range(1, 16):  # 所有15個欄位
                cell = merged_sheet.cell(row=row, column=col)
                cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
        
        # 實現垂直合併 - 整個工作表的A/B欄位垂直合併
        self._merge_ab_columns_globally(merged_sheet, current_row)
    
    def _create_parent_only_sheet(self, parent_data: List[Dict[str, Any]]):
        """創建只有父Chunk的工作表"""
        parent_sheet = self.workbook.create_sheet("父Chunks")
        
        # 設置標題行 - 父Chunk工作表不包含Sub Chunk欄位
        headers = [
            "原文", "正規化", "Parent Chunk", "層級", "Chunk ID", "父Chunk ID", 
            "索引", "大小", "是否表格", "標題層級", "標題文字", "頁碼", "檔名", "檔案類型"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = parent_sheet.cell(row=1, column=col, value=header)
            cell.font = cell.font.copy(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = self._get_header_fill()
        
        # 獲取原文和正規化內容（從第一個父chunk獲取）
        page_original = parent_data[0].get('original_content', '') if parent_data else ''
        page_normalized = parent_data[0].get('normalized_content', '') if parent_data else ''
        
        # 添加父chunk數據
        current_row = 2
        for i, parent in enumerate(parent_data):
            # 只有第一行設置原文和正規化內容，其他行留空（將垂直合併）
            if i == 0:
                parent_sheet.cell(row=current_row, column=1, value=page_original)  # 原文（第一行）
                parent_sheet.cell(row=current_row, column=2, value=page_normalized)  # 正規化（第一行）
            else:
                parent_sheet.cell(row=current_row, column=1, value="")  # 原文（留空，將合併）
                parent_sheet.cell(row=current_row, column=2, value="")  # 正規化（留空，將合併）
            
            parent_sheet.cell(row=current_row, column=3, value=parent['content'])  # Parent Chunk
            parent_sheet.cell(row=current_row, column=4, value="父層").font = parent_sheet.cell(row=current_row, column=4).font.copy(bold=True)
            parent_sheet.cell(row=current_row, column=5, value=parent['chunk_id'])
            parent_sheet.cell(row=current_row, column=6, value="")  # 父chunk沒有父ID
            parent_sheet.cell(row=current_row, column=7, value=parent.get('index', ''))
            parent_sheet.cell(row=current_row, column=8, value=parent['size'])
            parent_sheet.cell(row=current_row, column=9, value=parent['has_tables'])
            parent_sheet.cell(row=current_row, column=10, value=parent.get('header_level', ''))
            parent_sheet.cell(row=current_row, column=11, value=parent.get('header_text', ''))
            parent_sheet.cell(row=current_row, column=12, value=parent.get('page_number', ''))
            parent_sheet.cell(row=current_row, column=13, value=parent.get('file_name', ''))
            parent_sheet.cell(row=current_row, column=14, value=parent.get('file_type', ''))
            
            # 設置父chunk行格式
            self._format_parent_row(parent_sheet, current_row)
            current_row += 1
        
        # 調整列寬 - 父Chunk工作表
        self._adjust_parent_column_widths(parent_sheet)
        
        # 設置所有欄位垂直置頂對齊
        for row in range(2, current_row):
            for col in range(1, 15):  # 14個欄位
                cell = parent_sheet.cell(row=row, column=col)
                cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
        
        # 實現垂直合併 - 原文和正規化欄位
        if len(parent_data) > 1:
            # 合併原文欄位 (A欄)
            parent_sheet.merge_cells(f'A2:A{current_row - 1}')
            # 合併正規化欄位 (B欄)
            parent_sheet.merge_cells(f'B2:B{current_row - 1}')
    
    def _create_child_only_sheet(self, child_data: List[Dict[str, Any]]):
        """創建只有子Chunk的工作表"""
        child_sheet = self.workbook.create_sheet("子Chunks")
        
        # 設置標題行 - 子Chunk工作表不包含Parent Chunk欄位
        headers = [
            "原文", "正規化", "Sub Chunk", "層級", "Chunk ID", "父Chunk ID", 
            "索引", "大小", "是否表格", "標題層級", "標題文字", "頁碼", "檔名", "檔案類型"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = child_sheet.cell(row=1, column=col, value=header)
            cell.font = cell.font.copy(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = self._get_header_fill()
        
        # 獲取原文和正規化內容（從第一個子chunk獲取）
        page_original = child_data[0].get('original_content', '') if child_data else ''
        page_normalized = child_data[0].get('normalized_content', '') if child_data else ''
        
        # 添加子chunk數據
        current_row = 2
        for i, child in enumerate(child_data):
            # 只有第一行設置原文和正規化內容，其他行留空（將垂直合併）
            if i == 0:
                child_sheet.cell(row=current_row, column=1, value=page_original)  # 原文（第一行）
                child_sheet.cell(row=current_row, column=2, value=page_normalized)  # 正規化（第一行）
            else:
                child_sheet.cell(row=current_row, column=1, value="")  # 原文（留空，將合併）
                child_sheet.cell(row=current_row, column=2, value="")  # 正規化（留空，將合併）
            
            child_sheet.cell(row=current_row, column=3, value=child['content'])  # Sub Chunk
            child_sheet.cell(row=current_row, column=4, value="子層")
            child_sheet.cell(row=current_row, column=5, value=child['chunk_id'])
            child_sheet.cell(row=current_row, column=6, value=child['parent_chunk_id'])
            child_sheet.cell(row=current_row, column=7, value=child.get('child_index', ''))
            child_sheet.cell(row=current_row, column=8, value=child['size'])
            child_sheet.cell(row=current_row, column=9, value=child['is_table_chunk'])
            child_sheet.cell(row=current_row, column=10, value="")  # 子chunk沒有自己的標題層級
            child_sheet.cell(row=current_row, column=11, value=child.get('parent_header', ''))
            child_sheet.cell(row=current_row, column=12, value=child.get('page_number', ''))
            child_sheet.cell(row=current_row, column=13, value=child.get('file_name', ''))
            child_sheet.cell(row=current_row, column=14, value=child.get('file_type', ''))
            
            # 設置子chunk行格式
            self._format_child_row(child_sheet, current_row)
            current_row += 1
        
        # 調整列寬 - 子Chunk工作表
        self._adjust_child_column_widths(child_sheet)
        
        # 設置所有欄位垂直置頂對齊
        for row in range(2, current_row):
            for col in range(1, 15):  # 14個欄位
                cell = child_sheet.cell(row=row, column=col)
                cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
        
        # 實現垂直合併 - 原文和正規化欄位
        if len(child_data) > 1:
            # 合併原文欄位 (A欄)
            child_sheet.merge_cells(f'A2:A{current_row - 1}')
            # 合併正規化欄位 (B欄)
            child_sheet.merge_cells(f'B2:B{current_row - 1}')
    
    def _adjust_column_widths(self, sheet):
        """調整列寬 - 完整工作表（15欄）"""
        sheet.column_dimensions['A'].width = 50   # 原文
        sheet.column_dimensions['B'].width = 50   # 正規化
        sheet.column_dimensions['C'].width = 50   # Parent Chunk
        sheet.column_dimensions['D'].width = 50   # Sub Chunk
        sheet.column_dimensions['E'].width = 8    # 層級
        sheet.column_dimensions['F'].width = 20   # Chunk ID
        sheet.column_dimensions['G'].width = 20   # 父Chunk ID
        sheet.column_dimensions['H'].width = 8    # 索引
        sheet.column_dimensions['I'].width = 10   # 大小
        sheet.column_dimensions['J'].width = 12   # 是否表格
        sheet.column_dimensions['K'].width = 15   # 標題層級
        sheet.column_dimensions['L'].width = 30   # 標題文字
        sheet.column_dimensions['M'].width = 10   # 頁碼
        sheet.column_dimensions['N'].width = 30   # 檔名
        sheet.column_dimensions['O'].width = 15   # 檔案類型
    
    def _adjust_parent_column_widths(self, sheet):
        """調整列寬 - 父Chunk工作表（14欄）"""
        sheet.column_dimensions['A'].width = 50   # 原文
        sheet.column_dimensions['B'].width = 50   # 正規化
        sheet.column_dimensions['C'].width = 50   # Parent Chunk
        sheet.column_dimensions['D'].width = 8    # 層級
        sheet.column_dimensions['E'].width = 20   # Chunk ID
        sheet.column_dimensions['F'].width = 20   # 父Chunk ID
        sheet.column_dimensions['G'].width = 8    # 索引
        sheet.column_dimensions['H'].width = 10   # 大小
        sheet.column_dimensions['I'].width = 12   # 是否表格
        sheet.column_dimensions['J'].width = 15   # 標題層級
        sheet.column_dimensions['K'].width = 30   # 標題文字
        sheet.column_dimensions['L'].width = 10   # 頁碼
        sheet.column_dimensions['M'].width = 30   # 檔名
        sheet.column_dimensions['N'].width = 15   # 檔案類型
    
    def _adjust_child_column_widths(self, sheet):
        """調整列寬 - 子Chunk工作表（14欄）"""
        sheet.column_dimensions['A'].width = 50   # 原文
        sheet.column_dimensions['B'].width = 50   # 正規化
        sheet.column_dimensions['C'].width = 50   # Sub Chunk
        sheet.column_dimensions['D'].width = 8    # 層級
        sheet.column_dimensions['E'].width = 20   # Chunk ID
        sheet.column_dimensions['F'].width = 20   # 父Chunk ID
        sheet.column_dimensions['G'].width = 8    # 索引
        sheet.column_dimensions['H'].width = 10   # 大小
        sheet.column_dimensions['I'].width = 12   # 是否表格
        sheet.column_dimensions['J'].width = 15   # 標題層級
        sheet.column_dimensions['K'].width = 30   # 標題文字
        sheet.column_dimensions['L'].width = 10   # 頁碼
        sheet.column_dimensions['M'].width = 30   # 檔名
        sheet.column_dimensions['N'].width = 15   # 檔案類型
    
    def _merge_cells_by_page(self, sheet, page_parent_map, parent_child_map):
        """按頁面分組合併原文和正規化欄位"""
        current_row = 2
        
        for page_number in sorted(page_parent_map.keys()):
            page_parents = page_parent_map[page_number]
            
            # 計算該頁面的總行數
            page_rows = 0
            for parent in page_parents:
                parent_id = parent['chunk_id']
                children = parent_child_map.get(parent_id, [])
                page_rows += 1 + len(children)  # 1個父chunk + N個子chunks
            
            if page_rows > 1:
                # 合併原文欄位 (A欄)
                sheet.merge_cells(f'A{current_row}:A{current_row + page_rows - 1}')
                
                # 合併正規化欄位 (B欄)
                sheet.merge_cells(f'B{current_row}:B{current_row + page_rows - 1}')
            
            current_row += page_rows + 1  # +1 for separator row
    
    def _merge_ab_columns_globally(self, sheet, total_rows):
        """整個工作表的A/B欄位垂直合併"""
        if total_rows > 2:  # 確保有數據行
            # 合併原文欄位 (A欄) - 從第2行到最後一行
            sheet.merge_cells(f'A2:A{total_rows - 1}')
            
            # 合併正規化欄位 (B欄) - 從第2行到最後一行
            sheet.merge_cells(f'B2:B{total_rows - 1}')
    
    def _format_parent_row(self, sheet, row):
        """格式化父chunk行"""
        # 設置背景色
        from openpyxl.styles import PatternFill
        parent_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        
        for col in range(1, 16):  # 更新為15個欄位
            cell = sheet.cell(row=row, column=col)
            cell.fill = parent_fill
            cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
            if col == 5:  # 層級欄位
                cell.font = cell.font.copy(bold=True)
    
    def _format_child_row(self, sheet, row):
        """格式化子chunk行"""
        # 設置背景色
        from openpyxl.styles import PatternFill
        child_fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
        
        for col in range(1, 16):  # 更新為15個欄位
            cell = sheet.cell(row=row, column=col)
            cell.fill = child_fill
            cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
    
    def _get_header_fill(self):
        """獲取標題行填充樣式"""
        from openpyxl.styles import PatternFill
        return PatternFill(start_color="366092", end_color="366092", fill_type="solid")
